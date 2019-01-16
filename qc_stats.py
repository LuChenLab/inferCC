#!/usr/bin/env python3
# -*- coding:utf-8 -*-
u"""
Created by Zhang at 2019.01.12

整理cell ranger QC的结果
"""
import argparse as ap
import os
import sys

from lxml import etree
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

from pyecharts import Line, Boxplot, Page, Grid, Scatter


class QualityControl:
    u"""
    Extract from HTML, write into xlsx
    """

    name = [
        "Name",
        "Description",
        "Transcriptome",
        "Chemistry",
        "Cell Ranger Version",
        None,
        "Estimated Number of Cells",
        "Fraction Reads in Cells",
        "Mean Reads per Cell",
        "Median Genes per Cell",
        "Total Genes Detected",
        "Median UMI Counts per Cell",
        None,
        "Number of Reads",
        "Valid Barcodes",
        "Sequencing Saturation",
        "Q30 Bases in Barcode",
        "Q30 Bases in RNA Read",
        "Q30 Bases in UMI",
        None,
        "Reads Mapped to Genome",
        "Reads Mapped Confidently to Genome",
        "Reads Mapped Confidently to Intergenic Regions",
        "Reads Mapped Confidently to Intronic Regions",
        "Reads Mapped Confidently to Exonic Regions",
        "Reads Mapped Confidently to Transcriptome",
        "Reads Mapped Antisense to Gene",
    ]

    def __init__(self, *args, **kwargs):
        u"""
        init this class
        :param args: all positional args
        :param kwargs: all key value args
        """

        self.in_dir = os.path.abspath(kwargs["in_dir"])
        self.output = os.path.abspath(kwargs["output"])

        if not os.path.exists(self.in_dir) or not os.path.isdir(self.in_dir):
            raise ValueError("input dir")

        out_dir = os.path.dirname(self.output)
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)

        wb = load_workbook(kwargs["xlsx"])
        ws = wb[wb.sheetnames[0]]

        self.meta = {}
        for row in range(1, ws.max_row + 1):
            if ws[row][1].value and ws[row][1].value.startswith("2018"):
                self.meta[ws[row][1].value] = {
                    "owner": str(ws[row][9].value),
                    "tissue": str(ws[row][4].value)
                }

        self.data = self.__extract_all__()

    @staticmethod
    def __extract_single__(infile):
        u"""
        Extract all necessary information from HTML
        :return: {}
        """
        soup = etree.parse(infile, etree.HTMLParser())

        summary_cards = soup.xpath("//div[@class='summary_card']")

        data = {}
        for card in summary_cards[2:]:

            for tr in card.iter("tr"):
                tds = [x for x in tr.iter("td")]

                data[tds[0].text] = tds[1].text

        return data

    def __extract_all__(self):
        u"""
        extract all information from html files
        :return:
        """
        htmls = []
        for parent, _, files in os.walk(self.in_dir):
            for f in files:
                if f.endswith("web_summary.html"):
                    htmls.append(os.path.join(parent, f))

        data = []

        for i in tqdm(htmls):
            data.append(self.__extract_single__(i))

        return sorted(data, key=lambda x: x["Name"])

    def write_to_file(self):
        u"""
        Write to xlsx file
        :return:
        """
        wb = Workbook()

        ws = wb.active

        for idx, col in enumerate(self.name):
            ws.cell(row=1, column=idx + 3, value=col)

        for row, i in enumerate(self.data):

            for col, name in enumerate(self.name):
                if name:
                    ws.cell(row=row + 2, column=col + 3, value=i[name])
                else:
                    ws.cell(row=row + 2, column=col + 3, value=None)

            tmp = self.meta.get(i["Name"], None)
            if tmp:
                ws.cell(row=row + 2, column=2, value=tmp["tissue"])
                ws.cell(row=row + 2, column=1, value=tmp["owner"])

        wb.save(self.output)

    def plot(self):
        u"""
        Do a echarts plot to view all the conditions
        :return:
        """
        def custom_formatter(params):
            return params.value[3] + ": " + params.value[1]

        line = Line("", width="100%", height=800)

        attr = sorted([x["Name"] for x in self.data])
        names = [x for x in self.name[5:] if x]

        data = []

        for idx, i in enumerate(sorted(names)):
            if not i:
                continue
            tmp_value = []
            extra_name = []
            for x in self.data:
                tmp = self.meta.get(x["Name"], "NA")
                if isinstance(tmp, dict):
                    tmp.get("tissue", "NA")

                extra_name.append("%s (%s)" % (x["Name"], tmp))
                if x[i].endswith("%"):
                    tmp_value.append(float(x[i].replace("%", "")))
                else:
                    tmp_value.append(int(x[i].replace(",", "")))

            line.add(
                i,
                attr,
                tmp_value,
                is_datazoom_show=True,
                datazoom_type="both",
                datazoom_range=[0, 100],
                is_datazoom_extra_show=True,
                xaxis_rotate=30,
            )

            boxplot = Boxplot(i, width="60%", height=800)
            boxplot.add(
                i,
                [i],
                boxplot.prepare_data([tmp_value]),
                is_legend_show=False,
                tooltip_formatter=custom_formatter,
                is_datazoom_extra_show=True,
                datazoom_extra_range=[0, 100]
            )

            scatter = Scatter(width="60%", height=800)
            scatter.add(
                i,
                extra_name,
                tmp_value,
                extra_data=extra_name,
                extra_name=extra_name,
                is_legend_show=False,
                tooltip_formatter=custom_formatter,
                is_datazoom_extra_show=True,
                datazoom_extra_range=[0, 100],
                xaxis_type="category",
                xaxis_rotate=45
            )

            grid = Grid(
                width="100%",
                height=600,
            )

            grid.add(scatter, grid_right="25%", grid_bottom="25%")
            grid.add(boxplot, grid_left="80%", grid_bottom="25%")

            data.append(grid)

        grid = Page()
        grid.add(line)
        for i in data:
            grid.add(i)
        grid.render(self.output)


if __name__ == '__main__':
    parser = ap.ArgumentParser("Extract QC results")
    parser.add_argument(
        "-i", "--input",
        type=str,
        help="Path to directory contains the web_summary.html"
    )
    parser.add_argument(
        "-x",
        "--xlsx",
        type=str,
        help="Path to xlsx file"
    )
    parser.add_argument(
        "-p", "--plot",
        action="store_true",
        default=False
    )
    parser.add_argument(
        "-o", "--output",
        type=str,
        help="Path to output xlsx"
    )

    if len(sys.argv) <= 1:
        parser.print_help()
    else:
        args = parser.parse_args(sys.argv[1:])

        tmp = QualityControl(
            in_dir=args.input,
            output=args.output,
            xlsx=args.xlsx
        )

        if args.plot:
            tmp.plot()
        else:
            tmp.write_to_file()

