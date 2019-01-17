#!/usr/bin/env python3
# -*- coding:utf-8 -*-
u"""
Created by Zhang yiming at 2019.01.15

Run Seurat in batch
"""
import argparse as ap
import os
from glob import glob
import sys
import subprocess as sp
import multiprocessing as mp
import gzip
from tqdm import tqdm
from shutil import copytree, rmtree, copyfile
from openpyxl import load_workbook
from pypinyin import lazy_pinyin


__dir__ = os.path.dirname(os.path.abspath(__file__))
__script__ = os.path.join(__dir__, "run_seurat.R")
__script2__ = os.path.join(__dir__, "run_seurat2.R")
__static__ = os.path.join(__dir__, "static")


class RunSeurat:
    u"""
    Class to handle all issues
    """

    def __init__(self, args):
        u"""
        init this class
        :param args: argparse namespace
        """

        self.input = os.path.abspath(args.input)

        if not os.path.exists(self.input):
            raise ValueError("%s not exists" % self.input)

        self.output = os.path.abspath(args.output)

        if not os.path.exists(self.output):
            os.makedirs(self.output)

        self.xlsx = args.xlsx
        self.processes = args.processes

    @staticmethod
    def run_seurat_clustering_single(args):
        u"""
        run seurat
        :param args: {"input": , "output":}
        :return:
        """
        if not os.path.exists(args["output"]):
            os.makedirs(args["output"])

        cmds = [
            "Rscript",
            __script__,
            "-i {input} -o {output} -l {label}".format(**args)
        ]

        with open(os.devnull, "w") as w:
            try:
                sp.check_call(" ".join(cmds), shell=True, stdout=w, stderr=w)
            except sp.CalledProcessError:
                print(" ".join(cmds))

    def run_seurat_clustering(self):
        u"""
        Run seurat in batch
        :return:
        """
        args = []
        for i in os.listdir(self.input):
            args.append({
                "input": os.path.abspath(os.path.join(self.input, i)),
                "output": os.path.abspath(os.path.join(self.output, i)),
                "label": i
            })

        with mp.Pool(processes=self.processes) as pool:
            list(tqdm(pool.imap_unordered(self.run_seurat_clustering_single, args), total=len(args)))

    @staticmethod
    def run_seurat_comparison_single(args):
        cmds = [
            "Rscript",
            __script2__,
            "-c {cancer} -n {normal} -o {output} -l {label}".format(**args)
        ]

        with open(os.devnull, "w") as w:
            try:
                sp.check_call(" ".join(cmds), shell=True, stdout=w, stderr=w)
            except sp.CalledProcessError:
                print(" ".join(cmds))

    def run_seurat_comparison(self):
        u"""
        Run seurat comparison
        :return:
        """
        wb = load_workbook(self.xlsx)
        ws = wb[wb.sheetnames[0]]

        data = {}

        header = {
            "Name": None,
            "tissue_type": None,
            "status": None,
            "patient": None
        }
        for row_idx, row in enumerate(ws.rows):
            if row_idx == 0:
                for col_idx, col in enumerate(row):
                    if col.value in header.keys():
                        header[col.value] = col_idx
            elif not row[0].value:
                continue
            else:
                if row[header["status"]].value == 0:
                    continue

                if row[header["patient"]].value not in data.keys():
                    tmp = {}
                else:
                    tmp = data[row[header["patient"]].value]

                tmp[row[header["tissue_type"]].value.strip()] = row[header["Name"]].value

                data[str(row[header["patient"]].value.strip())] = tmp

        inputs = {}
        for i in os.listdir(self.input):
            inputs[i] = os.path.abspath(os.path.join(self.input, i))

        args = []
        for key, value in data.items():
            if len(value) < 2:
                continue

            key = "_".join(lazy_pinyin(key))

            out_dir = os.path.join(self.output, key)

            try:
                normal = value["normal"]

                for k, v in value.items():
                    if k == "normal":
                        continue

                    k = k.replace(" ", "_")

                    real_out_dir = os.path.join(out_dir, k)
                    if not os.path.exists(real_out_dir):
                        os.makedirs(real_out_dir)

                    args.append({
                        "normal": inputs[normal],
                        "cancer": inputs[v],
                        "label": k,
                        "output": real_out_dir
                    })

            except KeyError:
                continue

        with mp.Pool(processes=self.processes) as pool:
            list(tqdm(pool.imap_unordered(self.run_seurat_comparison_single, args), total=len(args)))


class Report:
    u"""
    Make html
    """

    def __init__(self, args):

        self.input = args.input

        if not os.path.exists(self.input):
            raise ValueError("%s not exists" % self.input)

        self.output = args.output

        if not os.path.exists(self.output):
            os.makedirs(self.output)
        self.xlsx = args.xlsx

    def generate_report_clustering(
            self,
            plots,
            columns
    ):
        u"""
        Generate report
        :return:
        """

        target = os.path.join(self.output, "static")

        if os.path.exists(target):
            rmtree(target)

        copytree(__static__, target)

        table = self.__generate_table_clustering__(columns)

        home = self.__generate_home_page__(table)

        with open(os.path.join(self.output, "index.html"), "w+") as w:
            w.write(home)

        out_html = os.path.join(self.output, "html")
        if not os.path.exists(out_html):
            os.makedirs(out_html)

        out_img = os.path.join(self.output, "img")
        if not os.path.exists(out_img):
            os.makedirs(out_img)

        for i in os.listdir(self.input):
            full_path = os.path.join(self.input, i)

            for j in os.listdir(full_path):
                if j.endswith("png"):

                    target = os.path.join(out_img, i)

                    if not os.path.exists(target):
                        os.makedirs(target)

                    copyfile(
                        os.path.join(full_path, j),
                        os.path.join(target, j)
                    )

            with open(os.path.join(out_html, i + ".html"), "w+") as w:
                w.write(self.__generate_single_page__(i, plots))

    def generate_report_comparison(
            self,
            plots,
            columns,
    ):
        u"""
        Generate report
        :return:
        """

        target = os.path.join(self.output, "static")

        if os.path.exists(target):
            rmtree(target)

        copytree(__static__, target)

        table = self.__generate_table_comparision__(columns)

        home = self.__generate_home_page__(table)

        with open(os.path.join(self.output, "index.html"), "w+") as w:
            w.write(home)

        out_html = os.path.join(self.output, "html")
        if not os.path.exists(out_html):
            os.makedirs(out_html)

        out_img = os.path.join(self.output, "img")
        if not os.path.exists(out_img):
            os.makedirs(out_img)

        files = glob(os.path.join(self.input, "*/*/*.png"))
        files_info = {}   # html_name: {order, image path, image alias}
        for f in files:
            patient = os.path.basename(os.path.dirname(os.path.dirname(f)))
            type_ = os.path.basename(os.path.dirname(f))

            key = "%s_%s" % (patient, type_)

            tmp = files_info.get(key, [])
            tmp.append({
                "name": os.path.basename(f),
                "source": f
            })
            files_info[key] = tmp

        for key, value in files_info.items():
            out_img_dir = os.path.join(out_img, key)

            if not os.path.exists(out_img_dir):
                os.makedirs(out_img_dir)

            for v in value:
                copyfile(
                    v["source"],
                    os.path.join(out_img_dir, v["name"])
                )

            with open(os.path.join(out_html, key + ".html"), "w+") as w:
                w.write(self.__generate_single_page__(indir=out_img_dir, plots=plots))

    def __generate_table_clustering__(self, columns):
        u"""
        Generate clustering home page table
        :return:
        """
        wb = load_workbook(self.xlsx)
        ws = wb[wb.sheetnames[0]]

        data = {"thead": "", "tbody": ""}

        for row in range(1, ws.max_row + 1):

            if not ws[row][0].value:
                continue

            tmp = "<tr>\n%s</tr>\n"
            tbody_row = []
            for col in columns:
                col -= 1

                if row == 1:
                    data["thead"] += "<th>%s</th>" % ws[row][col].value
                else:
                    if col == 0:
                        tbody_row.append(
                            '<td><a href="./html/{0}.html">{0}</a></td>'.format(ws[row][col].value)
                        )
                    else:
                        tbody_row.append(
                            "<td>%s</td>" % ws[row][col].value
                        )
            if tbody_row:
                data["tbody"] += tmp % "\n".join(tbody_row)

        return data

    def __generate_table_comparision__(self, columns):
        u"""

        :param table:
        :return:
        """

        wb = load_workbook(self.xlsx)
        ws = wb[wb.sheetnames[0]]

        meta_info = {}

        data = {"thead": "", "tbody": ""}

        header = {
            "Name": None,
            "tissue_type": None,
            "status": None,
            "patient": None
        }

        tmp_data = {}

        for row_idx, row in enumerate(ws.rows):
            if row_idx == 0:
                for col_idx, col in enumerate(row):
                    if col.value in header.keys():
                        header[col.value] = col_idx

                for i in columns:
                    data["thead"] += "<th>%s</th>" % row[i].value

            elif not row[0].value:
                continue
            else:
                if row[header["status"]].value == 0:
                    continue

                if row[header["patient"]].value not in meta_info.keys():
                    tmp = {}
                else:
                    tmp = meta_info[row[header["patient"]].value]

                tmp[row[header["tissue_type"]].value.strip()] = row[header["Name"]].value

                meta_info[str(row[header["patient"]].value.strip())] = tmp

                tmp_data[row[0].value] = []
                for i in columns:
                    tmp_data[row[0].value].append(row[i].value)

        for key, value in meta_info.items():
            if len(value) <= 1:
                continue

            try:
                normal = value["normal"]

                normal_row = []

                for i, j in enumerate(tmp_data[normal]):
                    if i == 0:
                        normal_row.append('<td rowspan="2">{link}</td>')
                    else:
                        normal_row.append('<td>%s</td>' % j)

                normal_row = "\n".join(normal_row)

                for k, v in value.items():

                    if k == "normal":
                        continue

                    link = '<a href="./html/{0}_{1}.html">{2}</a>'.format(
                        "_".join(lazy_pinyin(key)),
                        k.replace(" ", "_"),
                        key
                    )

                    cancer_row = []
                    for i in tmp_data[v][1:]:
                        cancer_row.append('<td>%s</td>' % i)

                    data["tbody"] += "<tr>%s</tr><tr>%s</tr>" % (
                        normal_row.format(**{"link": link}),
                        "\n".join(cancer_row)
                    )

            except KeyError:
                continue

        return data

    @staticmethod
    def __generate_home_page__(table):
        u"""
        :param columns: list of ints, to which columns are kept of xlsx
        :param table: {thead: "", tbody: "" }
        :return:
        """
        head = []
        jss = [
            "jquery-3.3.1.js",
            "jquery.dataTables.min.js",
            "dataTables.buttons.min.js",
            "buttons.bootstrap4.min.js",
            "jszip.min.js",
            "buttons.print.min.js",
            "buttons.html5.min.js"
        ]

        for i in sorted([x for x in os.listdir(__static__) if os.path.isfile(os.path.join(__static__, x))]):
            if i.startswith("."):
                continue
            elif i.endswith("css"):
                head.append(
                    '<link rel="stylesheet" href="./static/%s">' % i
                )

        for js in jss:
            head.append(
                '<script type="text/javascript" src="./static/%s"></script>' % js
            )

        head = "\n".join(head)

        table["head"] = head

        html = """
        <html>
            <head>
                <meta charset="utf-8">
                <title>Seurat report</title>
                %(head)s
            </head>
    
            <body>
                <div class="navbar navbar-expand-lg fixed-top navbar-dark bg-primary">
                    <div class="container">
                        <a href="./index.html" class="navbar-brand">Home</a> 
                    </div>
                </div>
                <div class="container">
                    <div class="bs-docs-section">
                        <div class="page-header">
                            <div class="row">
                                <div class="col-lg-12">
                                    <h3>统计信息</h3>
                                </div>
                            </div>
                        </div>
                    
                        <div class="row">
                            <div class="col-lg-12">
                                <div class="page-header">
                                    <h4 id="tables">测序质量</h4>
                                </div>
                                <div class="bs-component">
                                    <table id="table" class="table table-hover table-striped">
                                    <thead>
                                        <tr>
                                        %(thead)s
                                        </tr>
                                    </thead>
                                    <tbody>
                                        %(tbody)s
                                    </tbody>
                                    </table>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
                <script type="text/javascript">
                    $(document).ready(function() {
                        $("#table").DataTable({dom: "lBfrtip"});
                    } );
                </script>
                
                <div class="container">
                    <footer id="footer">
                        <div class="row">
                            <div class="col-lg-12">
                                <p>© 2019 Sichuan University (四川大学)</p>
                                <p>The Chenlinlab - State Key Laboratory of Biotherapy (生物治疗国家重点实验室)</p>
                            </div>
                        </div>
                    </footer>
                </div>
            </body>
        </html>
        """

        return html % table

    @staticmethod
    def __generate_single_page__(indir, plots):
        u"""
        :param plots: {fig name: fig alias}
        :return:
        """
        head = []
        for i in sorted([x for x in os.listdir(__static__) if os.path.isfile(os.path.join(__static__, x))]):
            if i.endswith("js"):
                head.append(
                    '<script type="text/javascript" src="../static/%s"></script>' % i
                )
            else:
                head.append(
                    '<link rel="stylesheet" href="../static/%s" >' % i
                )

        head = "\n".join(head)

        label = os.path.basename(indir)

        html = """
            <html>
                <head>
                    <meta charset="utf-8">
                    <title>%(label)s</title>
                    %(head)s
                </head>
    
                <body>
                    <div class="navbar navbar-expand-lg fixed-top navbar-dark bg-primary">
                      <div class="container">
                         <a href="../index.html" class="navbar-brand">Home</a> 
                      </div>
                    </div>
                    <div class="container">
                        <div class="page-header">
                            <div class="row">
                                <div class="col-lg-12">
                                    <h3>%(label)s</h3>
                                </div>
                            </div>
                        </div>
                        <div class="bs-docs-section">
                            <div class="row">
                                <div class="col-lg-12">
                                    <div class="bs-component">
                                        %(content)s
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                    <div class="container">
                        <footer id="footer">
                            <div class="row">
                                <div class="col-lg-12">
                                    <p>© 2019 Sichuan University (四川大学)</p>
                                    <p>The Chenlinlab - State Key Laboratory of Biotherapy (生物治疗国家重点实验室)</p>
                                </div>
                            </div>
                        </footer>
                    </div>
                </body>
            </html>
        """
        content = []
        for i in plots.keys():
            content.append(
                """
                <div class="bs-docs-section">
                    <div class="row">
                        <div class="col-lg-12">
                            <div class="page-header">
                                <h5>%(label)s</h5>
                            </div>
                            <div class="bs-component">
                                %(img)s
                            </div>
                        </div>
                    </div>
                </div>
                """ % ({
                    "img": '<img width="90%" height="800" src="../img/{0}/{1}" />'.format(label, i),
                    "label": plots[i]
                })
            )

        return html % ({
            "head": head,
            "label": label,
            "content": "\n".join(content)
        })


def __decompress__(args):
    f, outdir, lock = args

    sample_id = os.path.basename(
        os.path.dirname(
            os.path.dirname(
                os.path.dirname(f)
            )
        )
    )

    out_dir = os.path.join(outdir, sample_id)

    lock.acquire()
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    lock.release()

    basename = os.path.basename(f).replace(".gz", "")
    if basename == "features.tsv":
        basename = "genes.tsv"
    outfile = os.path.join(out_dir, basename)

    with open(outfile, "w+") as w:
        with gzip.open(f, "rb") as r:
            w.write(str(r.read().decode("utf-8")))


def decompress(indir, outdir, processes):
    u"""
    Compress data
    :param indir:
    :param outdir:
    :return:
    """
    input_files = glob(os.path.join(indir, "*/outs/filtered_feature_bc_matrix/*"))

    if not os.path.exists(outdir):
        os.makedirs(outdir)

    args = []
    m = mp.Manager()
    lock = m.Lock()
    for f in input_files:
        args.append([f, outdir, lock])

    with mp.Pool(processes) as pool:
        list(tqdm(pool.imap_unordered(__decompress__, args), total=len(args)))


if __name__ == '__main__':
    parser = ap.ArgumentParser("Run Seurat")
    parser.add_argument(
        "-i", "--input",
        help="Path to input directory",
        type=str,
        required=True
    )
    parser.add_argument(
        "-o", "--output",
        help="Path to output directory",
        type=str,
        required=True
    )
    parser.add_argument(
        "-p", "--processes",
        default=1,
        help="CPU usage",
        type=int,
    )
    parser.add_argument(
        "-x",
        "--xlsx",
        help="Path to meta info xlsx",
        required=True,
        type=str
    )
    parser.add_argument(
        "--decompress",
        action="store_true",
        default=False,
        help="Compress files"
    )
    parser.add_argument(
        "--clustering",
        action="store_true",
        default=False,
        help="Do clustering rather than comparision"
    )
    parser.add_argument(
        "--html",
        action="store_true",
        default=False,
        help="Collect results into output html"
    )

    if len(sys.argv) <= 1:
        parser.print_help()
    else:
        args = parser.parse_args(sys.argv[1:])

        if args.html:
            tmp = Report(args)

            if args.clustering:
                default_columns = [1, 2, 3, 4, 10, 11, 12, 13, 14, 15]

                default_plots = {
                    "VlnPlot.png": "Vln Plot",
                    "GenePlot.png": "Gene Plot",
                    "Variable.png": "Variable Genes",
                    "VizPCA.png": "Viz PCA",
                    "PCAPlot.png": "PCA Plot",
                    "PCHeatmap.png": "PCA Heatmap",
                    "JackStrawPlot.png": "Jack Straw Plot",
                    "PCElbowPlot.png": "PCElBowPlot",
                    "tSNE.png": "tSNE"
                }

                tmp.generate_report_clustering(plots=default_plots, columns=default_columns)
            else:
                default_columns = [4, 1, 2, 3, 10, 11, 12, 13, 14, 15]

                default_plots = {
                    "CCA.png": "CCA",
                    "MetageneBiorPlot.png": "Meta gene Bior Plot",
                    "DimHeatmap.png": "Dim Heatmap",
                    "AlignedCCAVlnPlot.png": "Aligned CCA Vln Plot",
                    "tSNE.png": "tSNE",
                    "feature.png": "feature",
                    "tSNE_with_cell_name.png": "tSNE with cell name",
                    "SpliceDotPlotGG.png": "Splice Dot Plot GG"
                }

                tmp.generate_report_comparison(plots=default_plots, columns=default_columns)

        elif args.decompress:
            decompress(
                args.input,
                args.output,
                args.processes
            )
        else:
            tmp = RunSeurat(args)

            if args.clustering:
                tmp.run_seurat_clustering()
            else:
                tmp.run_seurat_comparison()


